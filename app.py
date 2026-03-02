# -*- coding: utf-8 -*-
"""
app.py  –  Print Server
Chạy Flask trên 0.0.0.0 để các máy trong mạng LAN truy cập được.
"""

import io
import os
import sys
import uuid
import json
from datetime import datetime, timezone
from pathlib import Path

import requests as _requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from flask import (
    Flask, request, jsonify, render_template,
    send_from_directory, abort, send_file
)

# Thêm thư mục gốc vào sys.path để import core & error_handler
BASE_DIR = Path(__file__).parent
sys.path.insert(0, str(BASE_DIR))

from error_handler import log_error, log_info, log_warning
from scan_pdf import scan_pdf_for_orders
from sqlalchemy import func
from database import (
    init_db, get_session, UploadedFile, FileOrder, PrintJob, OrderPrint,
    PrintCheck, PrintCheckOrder
)

# ── Cấu hình ────────────────────────────────────────────────────────────────
UPLOAD_FOLDER        = BASE_DIR / "uploads"
EXCEL_FOLDER         = BASE_DIR / "excels"
JOB_LOG_FILE         = BASE_DIR / "logs" / "jobs.json"
PRINTER_ALIASES_FILE = BASE_DIR / "printer_aliases.json"
ALLOWED_EXT          = {"pdf"}
MAX_FILE_MB   = 50

UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
EXCEL_FOLDER.mkdir(parents=True, exist_ok=True)
JOB_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = str(UPLOAD_FOLDER)
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_MB * 1024 * 1024

# Khởi tạo database (tạo DB + bảng nếu chưa có)
try:
    init_db()
    log_info("Database initialized.")
except Exception as _db_err:
    log_error("init_db", _db_err)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _utcnow() -> datetime:
    """Trả về datetime UTC không có tzinfo (để lưu vào MySQL DATETIME)."""
    return datetime.now(timezone.utc).replace(tzinfo=None)

def _parse_note(note_raw: str | None) -> list:
    """Giải mã note JSON từ DB thành list dict. Trả về [] nếu rỗng hoặc lỗi."""
    if not note_raw:
        return []
    try:
        return json.loads(note_raw)
    except Exception:
        return []

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT


def load_jobs() -> list:
    if JOB_LOG_FILE.exists():
        try:
            with open(JOB_LOG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def save_jobs(jobs: list):
    with open(JOB_LOG_FILE, "w", encoding="utf-8") as f:
        json.dump(jobs, f, ensure_ascii=False, indent=2)


def add_job(filename: str, printer: str, status: str, message: str = "") -> dict:
    jobs = load_jobs()
    job = {
        "id":       str(uuid.uuid4())[:8],
        "filename": filename,
        "printer":  printer,
        "status":   status,   # success | error
        "message":  message,
        "time":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    jobs.insert(0, job)
    save_jobs(jobs[:200])     # giữ 200 bản ghi gần nhất
    return job


def get_printers() -> list[str]:
    """Lấy danh sách máy in trên Windows qua win32print."""
    try:
        import win32print
        printers = win32print.EnumPrinters(
            win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
        )
        return [p[2] for p in printers]
    except ImportError:
        log_warning("win32print không khả dụng – trả về danh sách rỗng.")
        return []
    except Exception as e:
        log_error("get_printers", e)
        return []


def get_default_printer() -> str:
    try:
        import win32print
        return win32print.GetDefaultPrinter()
    except Exception:
        return ""


def load_printer_aliases() -> dict:
    """Đọc alias: {"Beeprt BY-496": "Máy In Cắt", ...}"""
    if PRINTER_ALIASES_FILE.exists():
        try:
            with open(PRINTER_ALIASES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_printer_aliases(aliases: dict):
    with open(PRINTER_ALIASES_FILE, "w", encoding="utf-8") as f:
        json.dump(aliases, f, ensure_ascii=False, indent=2)


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# --- Danh sách máy in --------------------------------------------------------

@app.route("/api/printers")
def api_printers():
    printers = get_printers()
    default  = get_default_printer()
    aliases  = load_printer_aliases()
    printer_list = [
        {"id": p, "label": aliases.get(p, p)}
        for p in printers
    ]
    return jsonify({"printers": printer_list, "default": default})


# --- Cấu hình alias máy in ---------------------------------------------------

@app.route("/api/printer-aliases", methods=["GET"])
def api_get_aliases():
    aliases  = load_printer_aliases()
    printers = get_printers()
    result   = [{"id": p, "alias": aliases.get(p, "")} for p in printers]
    return jsonify({"aliases": result})


@app.route("/api/printer-aliases", methods=["POST"])
def api_set_alias():
    data    = request.get_json(force=True) or {}
    printer = data.get("printer", "").strip()
    alias   = data.get("alias", "").strip()
    if not printer:
        return jsonify({"ok": False, "error": "Thiếu tên máy in."}), 400
    aliases = load_printer_aliases()
    if alias:
        aliases[printer] = alias
    else:
        aliases.pop(printer, None)   # alias rỗng → xóa mapping
    save_printer_aliases(aliases)
    log_info(f"Alias máy in: '{printer}' → '{alias or '(đã xóa)'}'")
    return jsonify({"ok": True})


@app.route("/api/printer-aliases/<path:printer>", methods=["DELETE"])
def api_delete_alias(printer):
    aliases = load_printer_aliases()
    if printer in aliases:
        del aliases[printer]
        save_printer_aliases(aliases)
        log_info(f"Đã xóa alias máy in: '{printer}'")
    return jsonify({"ok": True})


# --- Upload file -------------------------------------------------------------

@app.route("/api/upload", methods=["POST"])
def api_upload():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Không có file được gửi lên."}), 400

    file = request.files["file"]
    if not file or file.filename == "":
        return jsonify({"ok": False, "error": "Tên file trống."}), 400

    if not allowed_file(file.filename):
        return jsonify({"ok": False, "error": "Chỉ chấp nhận file PDF."}), 400

    original_name = Path(file.filename).name
    timestamp     = _utcnow().strftime("%Y%m%d_%H%M%S")
    unique_name   = f"{timestamp}_{original_name}"
    dest          = UPLOAD_FOLDER / unique_name
    file.save(str(dest))

    file_size_kb = int(round(dest.stat().st_size / 1024, 0))
    upload_ip    = request.remote_addr
    now_utc      = _utcnow()

    # ── Quét PDF lấy danh sách đơn hàng ─────────────────────────
    scanned_orders    = []
    unrecognized_pages = []
    try:
        df_orders, unrecognized_pages = scan_pdf_for_orders(str(dest))
        if not df_orders.empty:
            scanned_orders = df_orders.to_dict("records")
    except Exception as e:
        log_error("api_upload.scan", e, {"filename": unique_name})

    # ── Ghi UploadedFile + FileOrder vào DB (1 transaction) ─────
    try:
        with get_session() as db:
            uf = UploadedFile(
                filename=unique_name,
                original_name=original_name,
                upload_time_utc=now_utc,
                upload_ip=upload_ip,
                file_size_kb=file_size_kb,
                note=json.dumps(unrecognized_pages, ensure_ascii=False) if unrecognized_pages else None,
            )
            db.add(uf)
            db.flush()  # lấy uf.id trước khi commit

            for order in scanned_orders:
                db.add(FileOrder(
                    uploaded_file_id    = uf.id,
                    filename            = unique_name,
                    order_sn            = order["order_sn"],
                    shop_name           = order.get("shop_name"),
                    platform            = order.get("platform"),
                    delivery_method     = order.get("delivery_method"),
                    delivery_method_raw = order.get("delivery_method_raw"),
                    page_number         = order.get("page"),
                ))
    except Exception as e:
        log_error("api_upload.db", e, {"filename": unique_name})

    log_info(f"Upload: {unique_name} từ {upload_ip} ({file_size_kb} KB) — {len(scanned_orders)} đơn")

    # ── Kiểm tra đơn trùng (đã in trước đó) ─────────────────────
    upload_warnings = []
    try:
        if scanned_orders:
            order_sns = [o["order_sn"] for o in scanned_orders]
            with get_session() as db:
                existing = db.query(OrderPrint).filter(
                    OrderPrint.order_sn.in_(order_sns)
                ).all()
                for op in existing:
                    upload_warnings.append({
                        "order_sn":        op.order_sn,
                        "shop_name":       op.shop_name,
                        "platform":        op.platform,
                        "delivery_method": op.delivery_method,
                        "print_count":     op.print_count,
                        "last_print_time": (
                            op.last_print_time_utc.strftime("%Y-%m-%d %H:%M:%S")
                            if op.last_print_time_utc else None
                        ),
                    })
    except Exception as e:
        log_error("api_upload.check_warnings", e, {"filename": unique_name})

    return jsonify({
        "ok":               True,
        "filename":         unique_name,
        "order_count":      len(scanned_orders),
        "upload_warnings":  upload_warnings,
        "unrecognized_pages": unrecognized_pages,
    })


# --- Danh sách file đã upload ------------------------------------------------

@app.route("/api/files")
def api_files():
    # Lấy note từ DB
    note_map: dict = {}
    try:
        with get_session() as db:
            for row in db.query(UploadedFile.filename, UploadedFile.note).all():
                note_map[row.filename] = row.note
    except Exception:
        pass

    files = []
    for p in sorted(UPLOAD_FOLDER.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if p.suffix.lower() == ".pdf":
            raw = note_map.get(p.name)
            note_items = []
            if raw:
                try:
                    note_items = json.loads(raw)
                except Exception:
                    pass
            files.append({
                "name":       p.name,
                "size_kb":    round(p.stat().st_size / 1024, 1),
                "time":       datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "note_items": note_items,
            })
    return jsonify({"files": files})


# --- Xóa file ----------------------------------------------------------------

@app.route("/api/files/<path:filename>", methods=["DELETE"])
def api_delete_file(filename):
    target = UPLOAD_FOLDER / filename
    if not target.exists():
        return jsonify({"ok": False, "error": "File không tồn tại."}), 404
    target.unlink()
    log_info(f"Đã xóa file: {filename}")
    return jsonify({"ok": True})


# --- Preview / download file -------------------------------------------------

@app.route("/api/files/<path:filename>", methods=["GET"])
def api_download_file(filename):
    return send_from_directory(str(UPLOAD_FOLDER), filename)


# --- Gửi lệnh in -------------------------------------------------------------

@app.route("/api/print/check", methods=["POST"])
def api_print_check():
    """Kiểm tra trước khi in: trả về cảnh báo nếu file/đơn đã in trước đó."""
    data     = request.get_json(force=True) or {}
    filename = data.get("filename", "").strip()

    if not filename:
        return jsonify({"ok": False, "error": "Thiếu tên file."}), 400

    filepath = UPLOAD_FOLDER / filename
    if not filepath.exists():
        return jsonify({"ok": False, "error": f"File không tồn tại: {filename}"}), 404

    result = {"ok": True, "has_warnings": False, "file_warnings": None, "order_warnings": [], "unrecognized_pages": []}

    # ── Kiểm tra file đã in chưa ─────────────────────────────────
    try:
        with get_session() as db:
            prev_jobs = (
                db.query(PrintJob)
                .filter(PrintJob.filename == filename, PrintJob.status == "success")
                .order_by(PrintJob.print_time_utc.desc())
                .all()
            )
            if prev_jobs:
                result["has_warnings"] = True
                latest = prev_jobs[0]
                result["file_warnings"] = {
                    "print_count":     len(prev_jobs),
                    "last_print_time": (
                        latest.print_time_utc.strftime("%Y-%m-%d %H:%M:%S")
                        if latest.print_time_utc else None
                    ),
                    "last_printer":    latest.printer_name,
                    "last_client_ip":  latest.client_ip,
                }
    except Exception as e:
        log_error("api_print_check.file", e)

    # ── Quét PDF → kiểm tra từng đơn đã in chưa ─────────────────
    try:
        # Ưu tiên dùng file_orders (đã scan lúc upload); fallback re-scan nếu file cũ
        with get_session() as db:
            file_order_rows = db.query(FileOrder).filter(FileOrder.filename == filename).all()
            # Chuyển sang dict ngay trong session để tránh detached instance error
            file_order_dicts = [
                {
                    "order_sn":        fo.order_sn,
                    "shop_name":       fo.shop_name,
                    "platform":        fo.platform,
                    "delivery_method": fo.delivery_method,
                    "page_number":     fo.page_number,
                }
                for fo in file_order_rows
            ]

        if file_order_dicts:
            order_sns = [fo["order_sn"] for fo in file_order_dicts]
            fo_map    = {fo["order_sn"]: fo for fo in file_order_dicts}
        else:
            # Backward compat: file upload trước khi có feature này
            df_orders, _unrecognized = scan_pdf_for_orders(str(filepath))
            result["unrecognized_pages"] = _unrecognized
            if not df_orders.empty:
                order_sns = df_orders["order_sn"].dropna().tolist()
                fo_map    = {row["order_sn"]: row for _, row in df_orders.iterrows()}
            else:
                order_sns = []
                fo_map    = {}

        if order_sns:
            with get_session() as db:
                existing = db.query(OrderPrint).filter(
                    OrderPrint.order_sn.in_(order_sns)
                ).all()
                if existing:
                    result["has_warnings"] = True
                    existing_map = {op.order_sn: op for op in existing}
                    for sn, fo in fo_map.items():
                        op = existing_map.get(sn)
                        if op:
                            page = fo.get("page_number") if isinstance(fo, dict) else fo.get("page")
                            result["order_warnings"].append({
                                "order_sn":        op.order_sn,
                                "shop_name":       op.shop_name,
                                "platform":        op.platform,
                                "delivery_method": op.delivery_method,
                                "page_number":     page,
                                "print_count":     op.print_count,
                                "last_print_time": (
                                    op.last_print_time_utc.strftime("%Y-%m-%d %H:%M:%S")
                                    if op.last_print_time_utc else None
                                ),
                            })
    except Exception as e:
        log_error("api_print_check.orders", e, {"filename": filename})

    # ── Lưu kết quả kiểm tra vào database ───────────────────────
    try:
        client_ip = request.remote_addr
        now_utc = _utcnow()
        
        with get_session() as db:
            # Tạo bản ghi PrintCheck
            print_check = PrintCheck(
                filename=filename,
                client_ip=client_ip,
                check_time_utc=now_utc,
                has_warnings=result["has_warnings"],
                file_printed_before=(result["file_warnings"] is not None),
                file_print_count=(
                    result["file_warnings"]["print_count"]
                    if result["file_warnings"] else None
                ),
                order_warnings_count=len(result["order_warnings"]),
                total_orders_in_file=len(fo_map) if 'fo_map' in locals() else 0,
            )
            db.add(print_check)
            db.flush()  # Lấy print_check.id

            # Lưu chi tiết các đơn có cảnh báo
            for order_warning in result["order_warnings"]:
                db.add(PrintCheckOrder(
                    print_check_id=print_check.id,
                    order_sn=order_warning["order_sn"],
                    shop_name=order_warning.get("shop_name"),
                    platform=order_warning.get("platform"),
                    delivery_method=order_warning.get("delivery_method"),
                    page_number=order_warning.get("page_number"),
                    print_count=order_warning.get("print_count", 0),
                    last_print_time_utc=(
                        datetime.strptime(order_warning["last_print_time"], "%Y-%m-%d %H:%M:%S")
                        if order_warning.get("last_print_time") else None
                    ),
                ))
        
        log_info(
            f"Kiểm tra in: {filename} từ {client_ip} - "
            f"{'CÓ CẢNH BÁO' if result['has_warnings'] else 'OK'} - "
            f"{len(result['order_warnings'])} đơn cảnh báo / {len(fo_map) if 'fo_map' in locals() else 0} đơn"
        )
    except Exception as e:
        log_error("api_print_check.save_db", e, {"filename": filename})

    return jsonify(result)


@app.route("/api/print", methods=["POST"])
def api_print():
    data           = request.get_json(force=True) or {}
    filename       = data.get("filename", "").strip()
    printer        = data.get("printer", "").strip()
    copies         = int(data.get("copies", 1))
    is_reprint     = bool(data.get("is_reprint", False))
    reprint_reason = data.get("reprint_reason", "").strip()
    client_ip      = request.remote_addr

    if not filename:
        return jsonify({"ok": False, "error": "Thiếu tên file."}), 400

    filepath = UPLOAD_FOLDER / filename
    if not filepath.exists():
        return jsonify({"ok": False, "error": f"File không tồn tại: {filename}"}), 404

    # ── Server-side validation: yêu cầu lý do nếu đã in trước đó ─
    try:
        with get_session() as db:
            prev_count = (
                db.query(PrintJob)
                .filter(PrintJob.filename == filename, PrintJob.status == "success")
                .count()
            )
        if prev_count > 0 and not is_reprint:
            return jsonify({
                "ok":                    False,
                "error":                 "File đã in trước đó. Vui lòng xác nhận lý do in lại.",
                "requires_reprint_reason": True,
            }), 409
    except Exception as e:
        log_error("api_print.check_prev", e)

    if is_reprint and not reprint_reason:
        return jsonify({"ok": False, "error": "Vui lòng nhập lý do in lại."}), 400

    # ── Lấy danh sách đơn hàng (từ DB hoặc fallback re-scan) ────
    orders_info = []
    try:
        with get_session() as db:
            file_order_rows = db.query(FileOrder).filter(FileOrder.filename == filename).all()
            # Chuyển sang dict ngay trong session để tránh detached instance error
            file_order_dicts = [
                {
                    "order_sn":            fo.order_sn,
                    "shop_name":           fo.shop_name,
                    "platform":            fo.platform or "unknown",
                    "delivery_method":     fo.delivery_method,
                    "delivery_method_raw": fo.delivery_method_raw or "",
                    "page":                fo.page_number,
                }
                for fo in file_order_rows
            ]
        if file_order_dicts:
            orders_info = file_order_dicts
        else:
            # Backward compat: file upload trước khi có feature này
            df_orders, _ = scan_pdf_for_orders(str(filepath))
            if not df_orders.empty:
                orders_info = df_orders.to_dict("records")
                
                # Lưu thành file excel - chỉ khi có đơn hàng
                try:
                    excel_filename = filepath.stem + ".xlsx"
                    excel_path = EXCEL_FOLDER / excel_filename
                    df_orders.to_excel(excel_path, index=False)
                    log_info(f"✅ Đã lưu thông tin đơn hàng vào excels/{excel_path.name}")
                except Exception as e:
                    log_error("save_excel", e, {"filename": filename})
            else:
                log_warning(f"Không tìm thấy đơn hàng nào trong {filename}")
    except Exception as e:
        log_error("api_print.orders", e, {"filename": filename})

    # ── Gửi lệnh in ──────────────────────────────────────────────
    try:
        sys.path.insert(0, str(BASE_DIR / "core"))
        from printing import print_pdf_printer

        success = True
        for _ in range(max(1, copies)):
            if not print_pdf_printer(str(filepath), printer or None):
                success = False
                break

        orders_summary = f"{len(orders_info)} đơn hàng" if orders_info else "Không xác định đơn hàng"
        now_utc        = _utcnow()
        status_str     = "success" if success else "error"

        # ── Ghi PrintJob vào DB ───────────────────────────────────
        try:
            with get_session() as db:
                db_job = PrintJob(
                    filename=filename,
                    printer_name=printer or "Default",
                    client_ip=client_ip,
                    copies=copies,
                    is_reprint=is_reprint,
                    reprint_reason=reprint_reason if is_reprint else None,
                    status=status_str,
                    print_time_utc=now_utc,
                )
                db.add(db_job)

                if success and orders_info:
                    for order in orders_info:
                        existing_op = db.query(OrderPrint).filter(
                            OrderPrint.order_sn == order["order_sn"]
                        ).first()
                        if existing_op:
                            existing_op.print_count         += 1
                            existing_op.last_print_time_utc  = now_utc
                            existing_op.filename             = filename
                        else:
                            db.add(OrderPrint(
                                filename=filename,
                                order_sn=order["order_sn"],
                                shop_name=order.get("shop_name"),
                                platform=order.get("platform", "unknown"),
                                delivery_method=order.get("delivery_method"),
                                delivery_method_raw=order.get("delivery_method_raw", ""),
                                page_number=order.get("page"),
                                print_count=1,
                                last_print_time_utc=now_utc,
                            ))
        except Exception as e:
            log_error("api_print.db", e)

        # Giữ jobs.json (backward compat)
        job = add_job(
            filename, printer or "Default", status_str,
            f"{copies} bản in - {orders_summary}" + (" [IN LẠI]" if is_reprint else "")
        )

        if success:
            log_info(f"In thành công: {filename} → {printer or 'Default'} x{copies} ({orders_summary})")
            return jsonify({"ok": True, "job": job, "orders": orders_info})
        else:
            msg = (
                "Gửi lệnh in thất bại. "
                "Gợi ý: Cài SumatraPDF (https://www.sumatrapdfreader.org) "
                "hoặc Adobe Acrobat Reader để in PDF tốt hơn."
            )
            return jsonify({"ok": False, "error": msg}), 500

    except Exception as e:
        log_error("api_print", e, {"filename": filename, "printer": printer})
        msg = str(e)
        add_job(filename, printer or "Default", "error", msg)
        return jsonify({"ok": False, "error": msg}), 500


# --- Lịch sử in (jobs.json - backward compat) --------------------------------

@app.route("/api/jobs")
def api_jobs():
    return jsonify({"jobs": load_jobs()})


# --- Lịch sử upload (DB) -----------------------------------------------------

@app.route("/api/files/history")
def api_files_history():
    page     = max(1, int(request.args.get("page", 1)))
    per_page = min(100, int(request.args.get("per_page", 20)))
    offset   = (page - 1) * per_page
    q        = request.args.get("q",  "").strip()
    ip       = request.args.get("ip", "").strip()
    try:
        with get_session() as db:
            qry = db.query(UploadedFile)
            if q:
                qry = qry.filter(UploadedFile.original_name.like(f"%{q}%"))
            if ip:
                qry = qry.filter(UploadedFile.upload_ip.like(f"%{ip}%"))
            total = qry.count()
            rows  = (
                qry
                .order_by(UploadedFile.upload_time_utc.desc())
                .offset(offset).limit(per_page).all()
            )
            # Đếm số đơn hàng theo filename (1 query)
            fnames = [r.filename for r in rows]
            order_counts: dict = {}
            print_stats: dict = {}
            if fnames:
                cnt_rows = (
                    db.query(FileOrder.filename, func.count(FileOrder.id))
                    .filter(FileOrder.filename.in_(fnames))
                    .group_by(FileOrder.filename)
                    .all()
                )
                order_counts = {fn: cnt for fn, cnt in cnt_rows}

                # Thống kê lệnh in theo filename
                pj_rows = (
                    db.query(
                        PrintJob.filename,
                        func.count(PrintJob.id).label("print_count"),
                        func.max(PrintJob.print_time_utc).label("last_print_time"),
                    )
                    .filter(PrintJob.filename.in_(fnames))
                    .group_by(PrintJob.filename)
                    .all()
                )
                # Lấy printer + status của lần in cuối cùng
                last_job_map: dict = {}
                if pj_rows:
                    last_times = {fn: lt for fn, _, lt in pj_rows if lt}
                    for fn, lt in last_times.items():
                        lj = (
                            db.query(PrintJob.printer_name, PrintJob.status)
                            .filter(PrintJob.filename == fn, PrintJob.print_time_utc == lt)
                            .first()
                        )
                        if lj:
                            last_job_map[fn] = {"printer": lj.printer_name, "status": lj.status}
                for fn, pc, lt in pj_rows:
                    print_stats[fn] = {
                        "print_count":     pc,
                        "last_print_time": lt.strftime("%Y-%m-%d %H:%M:%S") if lt else None,
                        "last_printer":    last_job_map.get(fn, {}).get("printer"),
                        "last_status":     last_job_map.get(fn, {}).get("status"),
                    }
            files = [
                {
                    "id":            r.id,
                    "filename":      r.filename,
                    "original_name": r.original_name,
                    "upload_time":   r.upload_time_utc.strftime("%Y-%m-%d %H:%M:%S") if r.upload_time_utc else None,
                    "upload_ip":     r.upload_ip,
                    "file_size_kb":  r.file_size_kb,
                    "order_count":   order_counts.get(r.filename, 0),
                    "note_items":    _parse_note(r.note),
                    **print_stats.get(r.filename, {
                        "print_count": 0, "last_print_time": None,
                        "last_printer": None, "last_status": None,
                    }),
                }
                for r in rows
            ]
        return jsonify({"ok": True, "files": files, "total": total, "page": page, "per_page": per_page})
    except Exception as e:
        log_error("api_files_history", e)
        return jsonify({"ok": False, "error": str(e)}), 500


# --- Lịch sử in (DB) ---------------------------------------------------------

@app.route("/api/print-history")
def api_print_history():
    page      = max(1, int(request.args.get("page", 1)))
    per_page  = min(100, int(request.args.get("per_page", 20)))
    offset    = (page - 1) * per_page
    q         = request.args.get("q",          "").strip()
    printer   = request.args.get("printer",    "").strip()
    ip        = request.args.get("ip",         "").strip()
    status    = request.args.get("status",     "").strip()
    is_reprint= request.args.get("is_reprint", "").strip()
    try:
        with get_session() as db:
            qry = db.query(PrintJob)
            if q:
                qry = qry.filter(PrintJob.filename.like(f"%{q}%"))
            if printer:
                qry = qry.filter(PrintJob.printer_name.like(f"%{printer}%"))
            if ip:
                qry = qry.filter(PrintJob.client_ip.like(f"%{ip}%"))
            if status:
                qry = qry.filter(PrintJob.status == status)
            if is_reprint in ("0", "1"):
                qry = qry.filter(PrintJob.is_reprint == (is_reprint == "1"))
            total = qry.count()
            rows  = (
                qry
                .order_by(PrintJob.print_time_utc.desc())
                .offset(offset).limit(per_page).all()
            )
            # Đếm số đơn hàng theo filename (1 query)
            fnames = [r.filename for r in rows]
            order_counts: dict = {}
            if fnames:
                cnt_rows = (
                    db.query(FileOrder.filename, func.count(FileOrder.id))
                    .filter(FileOrder.filename.in_(fnames))
                    .group_by(FileOrder.filename)
                    .all()
                )
                order_counts = {fn: cnt for fn, cnt in cnt_rows}
            jobs = [
                {
                    "id":             r.id,
                    "filename":       r.filename,
                    "printer_name":   r.printer_name,
                    "client_ip":      r.client_ip,
                    "copies":         r.copies,
                    "is_reprint":     r.is_reprint,
                    "reprint_reason": r.reprint_reason,
                    "status":         r.status,
                    "print_time":     r.print_time_utc.strftime("%Y-%m-%d %H:%M:%S") if r.print_time_utc else None,
                    "order_count":    order_counts.get(r.filename, 0),
                }
                for r in rows
            ]
        return jsonify({"ok": True, "jobs": jobs, "total": total, "page": page, "per_page": per_page})
    except Exception as e:
        log_error("api_print_history", e)
        return jsonify({"ok": False, "error": str(e)}), 500


# --- Đơn hàng trong 1 file (có trạng thái đã in / chưa in) ----------------------------

@app.route("/api/files/<path:filename>/orders")
def api_file_orders(filename):
    """Trả về danh sách đơn hàng trong file kèm trạng thái đã in / chưa in."""
    try:
        with get_session() as db:
            file_orders = (
                db.query(FileOrder)
                .filter(FileOrder.filename == filename)
                .order_by(FileOrder.page_number)
                .all()
            )
            if not file_orders:
                return jsonify({"ok": True, "orders": [], "total": 0,
                                "printed": 0, "unprinted": 0})

            # Lấy trạng thái in của tất cả đơn trong file (1 query)
            order_sns   = [fo.order_sn for fo in file_orders]
            printed_map = {
                op.order_sn: op
                for op in db.query(OrderPrint)
                           .filter(OrderPrint.order_sn.in_(order_sns))
                           .all()
            }

            orders = []
            for fo in file_orders:
                op = printed_map.get(fo.order_sn)
                orders.append({
                    "id":                fo.id,
                    "order_sn":          fo.order_sn,
                    "shop_name":         fo.shop_name,
                    "platform":          fo.platform,
                    "delivery_method":   fo.delivery_method,
                    "delivery_method_raw": fo.delivery_method_raw,
                    "page_number":       fo.page_number,
                    "printed":           op is not None,
                    "print_count":       op.print_count if op else 0,
                    "last_print_time":   (
                        op.last_print_time_utc.strftime("%Y-%m-%d %H:%M:%S")
                        if op and op.last_print_time_utc else None
                    ),
                })

        printed   = sum(1 for o in orders if o["printed"])
        unprinted = len(orders) - printed
        return jsonify({
            "ok":       True,
            "filename": filename,
            "orders":   orders,
            "total":    len(orders),
            "printed":  printed,
            "unprinted": unprinted,
        })
    except Exception as e:
        log_error("api_file_orders", e)
        return jsonify({"ok": False, "error": str(e)}), 500


# --- Lịch sử đơn hàng (DB) ---------------------------------------------------

@app.route("/api/orders/history")
def api_orders_history():
    page            = max(1, int(request.args.get("page", 1)))
    per_page        = min(100, int(request.args.get("per_page", 50)))
    order_sn        = request.args.get("order_sn",        "").strip()
    shop_name       = request.args.get("shop_name",       "").strip()
    platform        = request.args.get("platform",        "").strip()
    delivery_method = request.args.get("delivery_method", "").strip()
    offset          = (page - 1) * per_page
    try:
        with get_session() as db:
            qry = db.query(OrderPrint)
            if order_sn:
                qry = qry.filter(OrderPrint.order_sn.like(f"%{order_sn}%"))
            if shop_name:
                qry = qry.filter(OrderPrint.shop_name.like(f"%{shop_name}%"))
            if platform:
                qry = qry.filter(OrderPrint.platform == platform)
            if delivery_method:
                qry = qry.filter(OrderPrint.delivery_method == delivery_method)
            total = qry.count()
            rows  = qry.order_by(OrderPrint.id.desc()).offset(offset).limit(per_page).all()
            orders = [
                {
                    "id":              r.id,
                    "filename":        r.filename,
                    "order_sn":        r.order_sn,
                    "shop_name":       r.shop_name,
                    "platform":        r.platform,
                    "delivery_method": r.delivery_method,
                    "delivery_method_raw": r.delivery_method_raw,
                    "page_number":     r.page_number,
                    "print_count":     r.print_count,
                    "last_print_time": (
                        r.last_print_time_utc.strftime("%Y-%m-%d %H:%M:%S")
                        if r.last_print_time_utc else None
                    ),
                }
                for r in rows
            ]
        return jsonify({"ok": True, "orders": orders, "total": total, "page": page, "per_page": per_page})
    except Exception as e:
        log_error("api_orders_history", e)
        return jsonify({"ok": False, "error": str(e)}), 500


# --- Báo cáo Phiếu xuất kho cho 1 file --------------------------------------

OMS_BASE = os.environ.get("OMS_BASE_URL", "http://localhost:8000")


def _oms_get(path: str, **params):
    """GET request tới OMS, trả về dict JSON."""
    url = f"{OMS_BASE}{path}"
    resp = _requests.get(url, params=params, timeout=15)
    resp.raise_for_status()
    return resp.json()


def _oms_post(path: str, body):
    """POST request tới OMS, trả về dict JSON."""
    url = f"{OMS_BASE}{path}"
    resp = _requests.post(url, json=body, timeout=30)
    if not resp.ok:
        try:
            detail = resp.json()
        except Exception:
            detail = resp.text[:500]
        raise Exception(f"{resp.status_code} {resp.reason} — {detail}")
    return resp.json()


@app.route("/api/files/<path:filename>/report")
def api_file_report(filename):
    """
    Xuất Phiếu xuất kho dưới dạng file Excel (.xlsx) cho một file PDF đã upload.
    Luồng:
      1. Lấy danh sách đơn hàng từ DB, group theo shop_name
      2. Resolve shop_id từ OMS
      3. Gọi fetch-items để lấy items / model_quantity_purchased
      4. Gọi find-warehouse-sku để lấy warehouse_sku + warehouse_quantity
      5. Tổng hợp final_qty = model_quantity_purchased × warehouse_quantity
      6. Xuất Excel trả về cho browser download
    """
    try:
        # ── Bước 1: Lấy đơn hàng từ DB ─────────────────────────────────────
        with get_session() as db:
            rows = (
                db.query(FileOrder)
                .filter(FileOrder.filename == filename)
                .order_by(FileOrder.page_number)
                .all()
            )
            # Chuyển sang plain dict ngay trong session để tránh DetachedInstanceError
            file_orders = [
                {
                    "order_sn":  fo.order_sn,
                    "shop_name": fo.shop_name,
                    "platform":  fo.platform,
                    "delivery_method": fo.delivery_method,
                    "page_number": fo.page_number,
                }
                for fo in rows
            ]

        if not file_orders:
            return jsonify({"ok": False, "error": "File không có đơn hàng nào hoặc không tồn tại."}), 404

        # Group theo shop_name  →  {shop_name: [order_sn, ...]}
        from collections import defaultdict
        shop_orders: dict[str, list[str]] = defaultdict(list)
        for fo in file_orders:
            shop_orders[fo["shop_name"]].append(fo["order_sn"])

        # ── Bước 2: Resolve shop_id cho từng shop ───────────────────────────
        shop_id_map: dict[str, int] = {}  # shop_name → shop_id
        resolve_warnings: list[str] = []
        for shop_name in shop_orders:
            try:
                data = _oms_get("/api/shops/resolve-id", shop_name=shop_name)
                sid = data.get("shop_id") or data.get("id")
                if sid:
                    shop_id_map[shop_name] = int(sid)
                else:
                    raise Exception(f"Không tìm thấy shop_id trong response: {data}")
            except Exception as e:
                raise Exception(f"Lỗi khi resolve shop_id cho shop '{shop_name}': {e}")

        if not shop_id_map:
            return jsonify({"ok": False, "error": "Không resolve được shop_id cho bất kỳ shop nào.",
                            "warnings": resolve_warnings}), 502

        # ── Bước 3: fetch-items  →  map order_sn → items ────────────────────
        # items: [{item_id, model_id, model_quantity_purchased, item_name, model_name}]
        order_items: dict[str, list[dict]] = {}  # order_sn → list của items
        for shop_name, order_sn_list in shop_orders.items():
            sid = shop_id_map.get(shop_name)
            if sid is None:
                continue
            try:
                data = _oms_post("/api/orders/fetch-items", {
                    "order_sn_list": order_sn_list,
                    "shop_id": sid,
                })
                for order in data.get("orders", []):
                    order_items[order["order_sn"]] = order.get("items", [])
            except Exception as e:
                log_error("api_file_report:fetch-items", e)
                raise Exception(f"Lỗi khi gọi fetch-items cho shop '{shop_name}': {e}")

        # ── Bước 4: find-warehouse-sku  →  (item_id, model_id) → {warehouse_sku, warehouse_quantity} ──
        # Gom tất cả (shop_id, item_id, model_id) unique qua từng order
        sku_input_set: set[tuple] = set()
        for shop_name, order_sn_list in shop_orders.items():
            sid = shop_id_map.get(shop_name)
            if sid is None:
                continue
            for osn in order_sn_list:
                for item in order_items.get(osn, []):
                    sku_input_set.add((str(sid), str(item["item_id"]), str(item["model_id"])))

        sku_map: dict[tuple, dict] = {}  # (item_id, model_id) → {warehouse_sku, warehouse_quantity}
        if sku_input_set:
            sku_payload = [
                {"shop_id": sid, "item_id": iid, "model_id": mid}
                for sid, iid, mid in sku_input_set
            ]
            try:
                sku_data = _oms_post("/api/products/find-warehouse-sku", sku_payload)
                for entry in sku_data.get("found", []):
                    key = (str(entry["item_id"]), str(entry["model_id"]))
                    sku_map[key] = {
                        "warehouse_sku": entry.get("warehouse_sku"),
                        "warehouse_quantity": int(entry.get("warehouse_quantity")),
                        "note": None,
                    }

                not_found_sku = sku_data.get("not-found", [])
                if not_found_sku:
                    # report with note cảnh báo nhưng không fail toàn bộ report
                    for nf in not_found_sku:
                        key = (str(nf["item_id"]), str(nf["model_id"]))
                        sku_map[key] = {
                            "warehouse_sku": None,
                            "warehouse_quantity": 1,
                            "note": f"Không tìm thấy SKU kho cho shop_id={nf['shop_id']}, item_id={nf['item_id']}, model_id={nf['model_id']}"
                        }

            except Exception as e:
                log_error("api_file_report:find-warehouse-sku", e)
                raise Exception(f"Lỗi khi gọi find-warehouse-sku: {e}")

        # ── Bước 5: Tổng hợp theo warehouse_sku ────────────────────────────
        # sku_summary: key → {item_name, model_name, total_qty, order_count, note}
        # key = warehouse_sku (found) hoặc "__NF__{item_id}__{model_id}" (not found)
        sku_summary: dict[str, dict] = {}
        for order_sn, items in order_items.items():
            for item in items:
                qty_purchased = item.get("model_quantity_purchased") or 0
                key = (str(item["item_id"]), str(item["model_id"]))
                sku_info = sku_map.get(key)
                if not sku_info:
                    continue
                wsku  = sku_info["warehouse_sku"]
                wqty  = sku_info["warehouse_quantity"]
                note  = sku_info.get("note")
                final = qty_purchased * wqty
                # Dùng synthetic key cho not_found để không gộp dưới None
                summary_key = wsku if wsku else f"__NF__{item['item_id']}__{item['model_id']}"
                if summary_key not in sku_summary:
                    sku_summary[summary_key] = {
                        "item_name":   item.get("item_name", ""),
                        "model_name":  item.get("model_name", ""),
                        "total_qty":   0,
                        "order_count": 0,
                        "note":        note,
                    }
                sku_summary[summary_key]["total_qty"]   += final
                sku_summary[summary_key]["order_count"] += 1

        # ── Bước 6: Xuất Excel ──────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Phiếu xuất kho"

        # Style helpers
        header_font    = Font(bold=True, color="FFFFFF", size=11)
        header_fill    = PatternFill("solid", fgColor="1F4E79")
        center_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin_side      = Side(style="thin")
        thin_border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        alt_fill       = PatternFill("solid", fgColor="EBF3FB")

        # Tiêu đề bảng
        report_date = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value    = f"PHIẾU XUẤT KHO  —  {filename}  ({report_date})"
        title_cell.font     = Font(bold=True, size=13, color="1F4E79")
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 28

        # Header row
        headers = ["STT", "Mã SKU kho", "Tên sản phẩm", "Phân loại", "Số lượng", "Ghi chú"]
        col_widths = [6, 20, 50, 25, 12, 50]
        ws.append(headers)
        for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = width
        ws.row_dimensions[2].height = 22

        # Data rows
        note_font = Font(color="C00000", italic=True, size=9)
        for i, (summary_key, info) in enumerate(sorted(sku_summary.items()), start=1):
            row_num = i + 2
            # Hiển thị "—" thay vì synthetic key cho not_found
            wsku_display = summary_key if not summary_key.startswith("__NF__") else "—"
            note_text    = info.get("note") or ""
            ws.append([
                i,
                wsku_display,
                info["item_name"],
                info["model_name"],
                info["total_qty"],
                note_text,
            ])
            fill = alt_fill if i % 2 == 0 else None
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border    = thin_border
                cell.alignment = center_align if col_idx in (1, 5) else left_align
                if fill:
                    cell.fill = fill
            # Tô đỏ chữ đỏ cột Ghi chú nếu có nội dung
            if note_text:
                ws.cell(row=row_num, column=6).font = note_font
            ws.row_dimensions[row_num].height = 18

        # Freeze panes dưới header
        ws.freeze_panes = "A3"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = filename.replace("/", "_").replace("\\", "_")
        dl_name = f"Phieu_xuat_kho_{safe_name}.xlsx"

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=dl_name,
        )

    except Exception as e:
        log_error("api_file_report", e)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/files/<path:filename>/rescan", methods=["POST"])
def api_file_rescan(filename):
    """
    Quét lại file PDF đã upload để cập nhật danh sách đơn hàng.
    Luồng:
      1. Tìm file trong DB
      2. Quét lại PDF bằng scan_pdf_for_orders
      3. Xóa các FileOrder cũ
      4. Thêm FileOrder mới
      5. Cập nhật note (unrecognized_pages)
    """
    try:
        # Kiểm tra file có tồn tại trong DB không
        with get_session() as db:
            uf = db.query(UploadedFile).filter(UploadedFile.filename == filename).first()
            if not uf:
                return jsonify({"ok": False, "error": "File không tồn tại trong hệ thống."}), 404
        
        # Kiểm tra file vật lý có tồn tại không
        file_path = UPLOAD_FOLDER / filename
        if not file_path.exists():
            return jsonify({"ok": False, "error": "File vật lý không tồn tại."}), 404
        
        # Quét lại PDF
        scanned_orders = []
        unrecognized_pages = []
        try:
            df_orders, unrecognized_pages = scan_pdf_for_orders(str(file_path))
            if not df_orders.empty:
                scanned_orders = df_orders.to_dict("records")
        except Exception as e:
            log_error("api_file_rescan.scan", e, {"filename": filename})
            return jsonify({"ok": False, "error": f"Lỗi quét PDF: {str(e)}"}), 500
        
        # Cập nhật DB
        try:
            with get_session() as db:
                # Xóa các FileOrder cũ
                db.query(FileOrder).filter(FileOrder.filename == filename).delete()
                
                # Thêm FileOrder mới
                uf = db.query(UploadedFile).filter(UploadedFile.filename == filename).first()
                for order in scanned_orders:
                    db.add(FileOrder(
                        uploaded_file_id    = uf.id,
                        filename            = filename,
                        order_sn            = order["order_sn"],
                        shop_name           = order.get("shop_name"),
                        platform            = order.get("platform"),
                        delivery_method     = order.get("delivery_method"),
                        delivery_method_raw = order.get("delivery_method_raw"),
                        page_number         = order.get("page"),
                    ))
                
                # Cập nhật note
                uf.note = json.dumps(unrecognized_pages, ensure_ascii=False) if unrecognized_pages else None
                db.commit()
                
        except Exception as e:
            log_error("api_file_rescan.db", e, {"filename": filename})
            return jsonify({"ok": False, "error": f"Lỗi cập nhật database: {str(e)}"}), 500
        
        log_info(f"Rescan: {filename} — {len(scanned_orders)} đơn")
        
        return jsonify({
            "ok": True,
            "order_count": len(scanned_orders),
            "unrecognized_count": len(unrecognized_pages)
        })
        
    except Exception as e:
        log_error("api_file_rescan", e)
        return jsonify({"ok": False, "error": str(e)}), 500


# --- Thông tin máy chủ -------------------------------------------------------

@app.route("/api/info")
def api_info():
    import socket
    hostname  = socket.gethostname()
    local_ip  = socket.gethostbyname(hostname)
    return jsonify({
        "hostname": hostname,
        "ip":       local_ip,
        "port":     PORT,
    })


# ── Entry point ──────────────────────────────────────────────────────────────

PORT = int(os.environ.get("PRINT_SERVER_PORT", 5000))

if __name__ == "__main__":
    import socket
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = "127.0.0.1"

    print("=" * 60)
    print("  🖨️  PRINT SERVER đang chạy")
    print(f"  Local  : http://localhost:{PORT}")
    print(f"  Mạng LAN: http://{local_ip}:{PORT}")
    print("  Nhấn Ctrl+C để dừng")
    print("=" * 60)

    app.run(host="0.0.0.0", port=PORT, debug=False)
